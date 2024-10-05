import re
import openpyxl
from openpyxl.utils import get_column_letter

# Function to extract problem statements and hints from the text file
def extract_problems_and_hints(file_path):
    data = []
    
    # Define regex pattern to capture each problem's body and hint
    # It assumes "Problem:" starts the problem description, and "Hint:" starts the hint
    problem_pattern = re.compile(r"Problem:(.*?)Hint:\s*(.*?)(?=(Problem:|$))", re.DOTALL)
    
    with open(file_path, "r", encoding="utf-8") as file:
        content = file.read()

        # Use regex to find all problem blocks
        problems = problem_pattern.findall(content)
        
        for problem_body, hint, _ in problems:
            # Clean up extra whitespace and newlines
            problem_body = problem_body.strip()
            hint = hint.strip(".")
            
            # Append the cleaned problem and hint to the data list
            data.append([problem_body, hint])

    return data

# Function to write data to Excel using openpyxl
def write_to_excel(file_path, data):
    # Create a new workbook and active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the headers
    headers = ['Problem Statement', 'Hint']
    sheet.append(headers)

    # Set column widths for better readability
    sheet.column_dimensions[get_column_letter(1)].width = 80
    sheet.column_dimensions[get_column_letter(2)].width = 50

    # Append each row of problem statement and hint to the worksheet
    for row in data:
        sheet.append(row)

    # Save the workbook to the specified file
    workbook.save(file_path)
    print(f"Data has been successfully written to {file_path}")

# File path of the input .txt file containing all the questions
txt_file_path = "hard.txt"
# Output Excel file path
excel_output_file = "problems_hints_output.xlsx"

# Extract problems and hints from the .txt file
data = extract_problems_and_hints(txt_file_path)

# Write the extracted data to an Excel file using openpyxl
write_to_excel(excel_output_file, data)
