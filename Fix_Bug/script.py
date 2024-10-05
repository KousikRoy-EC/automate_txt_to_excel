import re
import openpyxl
from openpyxl.utils import get_column_letter

# Function to extract relevant information from the text file
def extract_data_from_txt(file_path):
    data = []
    current_language = ""
    current_code = ""
    current_error_hint = ""
    
    # Define regex patterns
    language_pattern = re.compile(r"### (C\+\+|JavaScript|Python|Java)")
    code_pattern = re.compile(r"```(cpp|javascript|python|java)\n(.*?)```", re.DOTALL)
    error_hint_pattern = re.compile(r"\*\*Error Hint:\*\* (.+)")
    problem_description_pattern = re.compile(r"\*\*Problem Statement:\*\*\s*(.*?)(?=\n###|```)", re.DOTALL)

    with open(file_path, 'r') as file:
        content = file.read()

        # Use regex to find problem blocks instead of splitting by question numbers
        problems = re.split(r"## Question \d+", content)
        for problem in problems[1:]:  # Skip the first empty split part

            # Extract the problem description between "Problem Statement" and code snippets
            problem_description_match = problem_description_pattern.search(problem)
            if problem_description_match:
                current_problem_description = problem_description_match.group(1).strip()
            else:
                # Fallback: extract first few lines if "Problem Statement" is not available
                description_lines = []
                for line in problem.splitlines():
                    if line.startswith('```'):
                        break
                    if not line.startswith('##') and not line.startswith('###'):
                        description_lines.append(line.strip())
                current_problem_description = "\n".join(description_lines).strip()

            # Find all code snippets in the problem
            codes = code_pattern.findall(problem)
            error_hints = error_hint_pattern.findall(problem)

            for i, (lang_code, code_snippet) in enumerate(codes):
                current_language = lang_code
                current_code = code_snippet.strip()
                current_error_hint = error_hints[i] if i < len(error_hints) else ""

                # Append data to list
                data.append([current_language, f"{current_problem_description}\n\n{current_code}", current_error_hint])

    return data

# Function to write data to Excel using openpyxl
def write_to_excel(file_path, data):
    # Create a new workbook and active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the headers
    headers = ['Language', 'Problem Description + Code', 'Error Hint']
    sheet.append(headers)

    # Set column widths
    sheet.column_dimensions[get_column_letter(1)].width = 20
    sheet.column_dimensions[get_column_letter(2)].width = 80
    sheet.column_dimensions[get_column_letter(3)].width = 40

    # Append the data to the worksheet
    for row in data:
        sheet.append(row)

    # Save the workbook
    workbook.save(file_path)
    print(f"Data has been successfully written to {file_path}")

# File path of the .txt file
txt_file_path = "hard.txt"
# Output Excel file path
excel_output_file = "output_problems_openpyxl.xlsx"

# Extract data from the .txt file
data = extract_data_from_txt(txt_file_path)

# Write the data to an Excel file using openpyxl
write_to_excel(excel_output_file, data)
