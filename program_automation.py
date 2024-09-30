import openpyxl
from openpyxl.styles import Alignment

input_file = "dsa_questions.txt"
output_file = "dsa_questions.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DSA Questions"

ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 70
ws.column_dimensions["C"].width = 30

row_height = 50

alignment = Alignment(wrap_text=True)

with open("dsa_questions.txt", "r", encoding="utf-8") as file:
    lines = file.readlines()

languages = ["Java", "C++", "Python", "JavaScript"]
row = 1

i = 0
while i < len(lines):
    if lines[i].startswith("##"):
        problem_desc = []
        code_snippets = []
        error_hints = []

        # Parse the problem description
        while not lines[i].startswith("```") and i < len(lines):
            problem_desc.append(lines[i].strip())
            i += 1

        problem_desc = problem_desc[2:] 

        for lang in languages:
            code_snippet = []
            error_hint = None

            while i < len(lines) and not lines[i].startswith("```"):
                i += 1

            if i < len(lines) and lines[i].startswith("```"):
                i += 1 

                while i < len(lines) and not lines[i].startswith("```"):
                    code_snippet.append(lines[i].strip())
                    i += 1

                i += 1 

                while i < len(lines) and not lines[i].startswith("Error hint"):
                    i += 1

                if i < len(lines) and lines[i].startswith("Error hint"):
                    error_hint = lines[i].split(":")[1].strip()
                    i += 1

            ws.cell(row=row, column=1, value=lang).alignment = alignment
            ws.cell(row=row, column=2, value="\n".join(problem_desc + [""] + code_snippet)).alignment = alignment
            ws.cell(row=row, column=3, value=error_hint).alignment = alignment

            ws.row_dimensions[row].height = row_height
            row += 1

        i += 1

wb.save(output_file)

print(f"Excel file created successfully: {output_file}")